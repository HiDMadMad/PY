from PIL import Image
import openpyxl
from tqdm import tqdm

img = Image.open("files\\image_for_handout.jpg")
width, height = img.size

img_excel = openpyxl.Workbook()
first_sheet = img_excel.active
first_sheet.title = "image"

#------------------------------------------------------------------
'''
alphabetical = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
def cells_name(chars):
    list_of_cells = []
    for char in chars:
        list_of_cells.append(char)

    for char1 in chars:
        for char2 in chars:
            list_of_cells.append(char1+char2)

    for char1 in chars:
        for char2 in chars:
            for char3 in chars:
                list_of_cells.append(char1+char2+char3)
    return list_of_cells
cells = cells_name(alphabetical)  # first_sheet[cells[0]] --> 'A1'
'''
#------------------------------------------------------------------
def RGB_to_Hex(image):
    def rgb2hex(r, g, b):
        return 'FF{:02x}{:02x}{:02x}'.format(r, g, b)

    image_pixels_hex = []
    pixels = image.convert('RGB').load()

    for y in tqdm(range(0, height), desc="processing "):
            for x in range(0, width):
                r, g, b= pixels[x, y]
                image_pixels_hex.append(rgb2hex(r, g, b))  # 'x = %s, y = %s, hex = %s' % (x, y, rgb2hex(r, g, b))
    return image_pixels_hex
#------------------------------------------------------------------

pixels_colors = RGB_to_Hex(img)

for y in tqdm(range(1, height+1), desc="filling cells "):
    for x in range(1, width+1):
        cell_index = ((y-1)*width)+(x-1)
        first_sheet.cell(y, x).fill = openpyxl.styles.PatternFill(start_color=pixels_colors[cell_index], end_color=pixels_colors[cell_index], fill_type="solid")

img_excel.save("files\\complete_img_in_excel.xlsx")
print("\nyour artwork is ready sir!!!\n\n\n")
img_excel.close()
#MadMad_56